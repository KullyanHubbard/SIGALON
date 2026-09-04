"""Pembangun berkas ekspor data penduduk (.xlsx dan .csv).

Bentuk kolom dan tampilannya disamakan dengan `docs/template-data-penduduk.xlsx`
supaya hasil unduhan pengurus bisa langsung dibaca rapi di Excel atau diimpor
kembali ke sistem jika dibutuhkan.
"""

import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.data.impor_excel import BARIS_HEADER, KOLOM, NAMA_SHEET, PILIHAN
from app.schemas.penduduk import Penduduk

_ALAMAT = {
    "jalan",
    "rt",
    "rw",
    "desa",
    "kecamatan",
    "kabupaten",
    "provinsi",
    "kodePos",
}


def _nilai(p: Penduduk, field: str) -> str:
    """Ekstrak nilai string dari objek Penduduk berdasarkan nama field."""
    if field in _ALAMAT:
        return getattr(p.alamat, field, "") or ""
    val = getattr(p, field, "")
    return "" if val is None else str(val)


def bikin_excel(daftar: list[Penduduk], judul: str = "DATA PENDUDUK") -> bytes:
    """Buat file Excel (.xlsx) berisi data penduduk dengan styling dokumen resmi."""
    wb = Workbook()
    ws = wb.active
    ws.title = NAMA_SHEET

    tipis = Side(style="thin", color="B7B7B7")
    bingkai = Border(left=tipis, right=tipis, top=tipis, bottom=tipis)

    # Baris 1: Judul Laporan Tergabung
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(KOLOM))
    c_judul = ws.cell(row=1, column=1, value=judul)
    c_judul.fill = PatternFill("solid", fgColor="1F4E78")
    c_judul.font = Font(color="FFFFFF", bold=True, size=13)
    c_judul.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Baris 2: Header Kolom
    for i, (_, label, lebar) in enumerate(KOLOM, start=1):
        c = ws.cell(row=BARIS_HEADER, column=i, value=label)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
        c.font = Font(bold=True)
        c.border = bingkai
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = lebar
    ws.row_dimensions[BARIS_HEADER].height = 30
    ws.freeze_panes = f"A{BARIS_HEADER + 1}"

    # Baris 3+: Data Warga
    baris_awal = BARIS_HEADER + 1
    for r, p in enumerate(daftar, start=baris_awal):
        for i, (field, _, _) in enumerate(KOLOM, start=1):
            val = _nilai(p, field)
            c = ws.cell(row=r, column=i, value=val)
            c.border = bingkai

    baris_akhir = max(baris_awal, baris_awal + len(daftar) - 1)

    # Pasang dropdown validasi data untuk kolom-kolom pilihan
    if len(daftar) > 0:
        for i, (field, _, _) in enumerate(KOLOM, start=1):
            if field not in PILIHAN:
                continue
            opsi = PILIHAN[field]
            huruf = get_column_letter(i)
            dv = DataValidation(
                type="list",
                formula1=f'"{",".join(opsi)}"',
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Nilai tidak valid",
                error="Pilih salah satu: " + ", ".join(opsi),
            )
            ws.add_data_validation(dv)
            dv.add(f"{huruf}{baris_awal}:{huruf}{baris_akhir}")

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def bikin_csv(daftar: list[Penduduk]) -> bytes:
    """Buat file CSV ber-encoding UTF-8 BOM agar rapi saat dibuka di Microsoft Excel."""
    sio = io.StringIO()
    writer = csv.writer(sio, delimiter=",", quoting=csv.QUOTE_MINIMAL)

    # Header
    writer.writerow([label for _, label, _ in KOLOM])

    # Baris Data
    for p in daftar:
        writer.writerow([_nilai(p, field) for field, _, _ in KOLOM])

    # Tambahkan UTF-8 BOM (\xef\xbb\xbf) agar Excel langsung mengenali encoding UTF-8
    return b"\xef\xbb\xbf" + sio.getvalue().encode("utf-8")
